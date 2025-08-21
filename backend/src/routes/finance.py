# backend/src/routes/finance.py
import os
import json
import base64
import uuid
import logging
from datetime import datetime, timedelta, date
from typing import Any, Dict, List, Optional
from io import BytesIO

from flask import Blueprint, request, jsonify, Response
from flask_jwt_extended import verify_jwt_in_request, get_jwt, get_jwt_identity
from werkzeug.utils import secure_filename

log = logging.getLogger(__name__)
finance_bp = Blueprint("finance", __name__)

# ============================ Config ============================
_VALID_TYPES = {"entrada", "saida", "despesa"}
_COLL = os.getenv("FINANCE_COLL", "finance_transactions")
_JSON_PATH = os.getenv("FINANCE_JSON_PATH") or "/tmp/finance_transactions.json"
_USE_FIRESTORE = (os.getenv("USE_FIRESTORE", "true").lower() in ("1", "true", "yes"))

# ============================ Auth leve =========================
def _require_admin():
    """JWT opcional; permite role/roles=admin; senão 401 (sem 500)."""
    try:
        verify_jwt_in_request(optional=True)
        claims = get_jwt() or {}
        roles = claims.get("roles") or claims.get("role") or []
        if isinstance(roles, str):
            roles = [roles]
        if "admin" in roles:
            return None
    except Exception as e:
        log.debug("[finance][auth] sem JWT válido: %s", e)
    return jsonify({"message": "unauthorized"}), 401


# ============================ Utils =============================
def _now_iso() -> str:
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


def _iso_date_only(d: Any) -> str:
    """Normaliza p/ YYYY-MM-DD (aceita date/datetime/ISO/BR)."""
    if isinstance(d, (datetime, date)):
        return d.strftime("%Y-%m-%d")
    s = str(d or "").strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    try:  # 2025-08-20T12:34:56
        return datetime.strptime(s[:19], "%Y-%m-%dT%H:%M:%S").strftime("%Y-%m-%d")
    except Exception:
        return ""


def _to_number(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        try:
            return float(v)
        except Exception:
            return 0.0


def _month_bounds(ym: str) -> Optional[tuple[str, str]]:
    if not ym or len(ym) != 7 or "-" not in ym:
        return None
    y, m = map(int, ym.split("-"))
    start = datetime(y, m, 1)
    end = datetime(y + (m == 12), (m % 12) + 1, 1) - timedelta(days=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def _decode_b64_json(b64) -> Optional[dict]:
    try:
        return json.loads(base64.b64decode(b64).decode("utf-8"))
    except Exception:
        return None


# ============================ Firestore =========================
_fs_client = None
_fs_ready: Optional[bool] = None


def _get_firestore():
    """Tenta inicializar Firestore. Se falhar, usa fallback JSON sem quebrar."""
    global _fs_client, _fs_ready
    if not _USE_FIRESTORE:
        _fs_ready = False
        return None
    if _fs_ready is not None:
        return _fs_client if _fs_ready else None

    try:
        from google.cloud import firestore  # type: ignore
        from google.oauth2 import service_account  # type: ignore
    except Exception as e:
        log.warning("[finance] firestore lib ausente: %s (fallback JSON)", e)
        _fs_ready = False
        return None

    try:
        if os.getenv("FIREBASE_CREDENTIALS_B64"):
            info = _decode_b64_json(os.getenv("FIREBASE_CREDENTIALS_B64"))
            creds = service_account.Credentials.from_service_account_info(info)
            project = (info or {}).get("project_id") or os.getenv("FIREBASE_PROJECT_ID")
            _fs_client = firestore.Client(credentials=creds, project=project)
        elif os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON"):
            info = json.loads(os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON"))
            creds = service_account.Credentials.from_service_account_info(info)
            _fs_client = firestore.Client(credentials=creds, project=info.get("project_id"))
        else:
            _fs_client = firestore.Client()
        _fs_ready = True
        return _fs_client
    except Exception as e:
        log.warning("[finance] Firestore indisponível: %s (fallback JSON)", e)
        _fs_ready = False
        _fs_client = None
        return None


# ============================ Fallback JSON ======================
def _json_load() -> List[Dict[str, Any]]:
    try:
        if not os.path.exists(_JSON_PATH):
            return []
        with open(_JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception as e:
        log.exception("[finance] erro lendo JSON: %s", e)
        return []


def _json_save(items: List[Dict[str, Any]]) -> None:
    try:
        os.makedirs(os.path.dirname(_JSON_PATH), exist_ok=True)
        with open(_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False)
    except Exception as e:
        log.exception("[finance] erro salvando JSON: %s", e)


# ============================ I/O helpers =======================
def _doc_to_json(d: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(d)
    out["amount"] = _to_number(out.get("amount"))
    if out.get("date"):
        n = _iso_date_only(out["date"])
        if n:
            out["date"] = n
    if not out.get("id"):
        out["id"] = str(uuid.uuid4())
    return out


def _fs_list(ttype: str, action_id: str, month: str) -> List[Dict[str, Any]]:
    client = _get_firestore()
    items: List[Dict[str, Any]] = []
    if not client:
        items = _json_load()
    else:
        try:
            coll = client.collection(_COLL)
            q = coll
            if ttype in _VALID_TYPES:
                q = q.where("type", "==", ttype)
            if action_id:
                q = q.where("action_id", "==", action_id)
            if month:
                rng = _month_bounds(month)
                if rng:
                    a, b = rng
                    q = q.where("date", ">=", a).where("date", "<=", b)
            for snap in q.stream():
                d = snap.to_dict()
                d["id"] = snap.id
                items.append(_doc_to_json(d))
        except Exception as e:
            log.exception("[finance][fs_list] %s", e)
            items = _json_load()

    if ttype in _VALID_TYPES:
        items = [x for x in items if (x.get("type") or "") == ttype]
    if action_id:
        items = [x for x in items if str(x.get("action_id") or "") == action_id]
    if month:
        rng = _month_bounds(month)
        if rng:
            a, b = rng
            items = [x for x in items if a <= (x.get("date") or "") <= b]

    items.sort(key=lambda x: ((x.get("date") or ""), (x.get("created_at") or "")), reverse=True)
    return items


def _fs_create(doc: Dict[str, Any]) -> Dict[str, Any]:
    client = _get_firestore()
    if not client:
        rows = _json_load()
        rows.append(_doc_to_json(doc))
        _json_save(rows)
        return doc
    try:
        client.collection(_COLL).document(doc["id"]).set(doc)
        return doc
    except Exception as e:
        log.exception("[finance][fs_create] %s", e)
        rows = _json_load()
        rows.append(_doc_to_json(doc))
        _json_save(rows)
        return doc


def _fs_update(txid: str, upd: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    client = _get_firestore()
    if not client:
        rows = _json_load()
        for i, it in enumerate(rows):
            if str(it.get("id")) == str(txid):
                rows[i].update(upd)
                _json_save(rows)
                return rows[i]
        return None
    try:
        from google.cloud.exceptions import NotFound  # type: ignore
        ref = client.collection(_COLL).document(txid)
        snap = ref.get()
        if not snap.exists:
            return None
        ref.update(upd)
        out = ref.get().to_dict()
        out["id"] = txid
        return out
    except Exception as e:
        log.exception("[finance][fs_update] %s", e)
        return None


def _fs_delete(txid: str) -> bool:
    client = _get_firestore()
    if not client:
        rows = _json_load()
        _json_save([x for x in rows if str(x.get("id")) != str(txid)])
        return True
    try:
        ref = client.collection(_COLL).document(txid)
        if ref.get().exists:
            ref.delete()
        return True
    except Exception as e:
        log.exception("[finance][fs_delete] %s", e)
        rows = _json_load()
        _json_save([x for x in rows if str(x.get("id")) != str(txid)])
        return True


# ============================ CORS extra ========================
@finance_bp.after_request
def _add_cors_headers(resp: Response):
    origin = request.headers.get("Origin") or "*"
    resp.headers.setdefault("Access-Control-Allow-Origin", origin)
    resp.headers.setdefault("Vary", "Origin")
    resp.headers.setdefault("Access-Control-Allow-Methods", "GET,POST,PUT,PATCH,DELETE,OPTIONS")
    resp.headers.setdefault("Access-Control-Allow-Headers", "Authorization, Content-Type")
    resp.headers.setdefault("Access-Control-Expose-Headers", "Authorization, Content-Type")
    resp.headers.setdefault("Access-Control-Max-Age", "86400")
    return resp


@finance_bp.route("/transactions", methods=["OPTIONS"])
@finance_bp.route("/transactions/<path:_id>", methods=["OPTIONS"])
@finance_bp.route("/contas-pagar", methods=["OPTIONS"])
@finance_bp.route("/contas-pagar/<path:_id>", methods=["OPTIONS"])
@finance_bp.route("/contas-receber", methods=["OPTIONS"])
@finance_bp.route("/contas-receber/<path:_id>", methods=["OPTIONS"])
@finance_bp.route("/contas-pagar/import", methods=["OPTIONS"])
@finance_bp.route("/contas-receber/import", methods=["OPTIONS"])
def _preflight(_id=None):
    return jsonify(ok=True), 204


# ============================ Payload maps ======================
def _map_pagar(data: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "type": "saida",
        "date": _iso_date_only(data.get("vencimento") or data.get("date") or data.get("data")),
        "amount": abs(_to_number(data.get("valor") if data.get("valor") is not None else data.get("amount"))),
        "category": (data.get("categoria") or data.get("category") or "conta_pagar").strip(),
        "notes": (data.get("descricao") or data.get("notes") or "").strip(),
        "documento": (data.get("documento") or "").strip() or None,
        "dataPagamento": _iso_date_only(data.get("dataPagamento")) or None,
        "valorPago": _to_number(data.get("valorPago")),
        "action_id": (data.get("action_id") or "").strip() or None,
    }


def _map_receber(data: Dict[str, Any]) -> Dict[str, Any]:
    notes = data.get("descricao") or data.get("cliente") or data.get("notaFiscal") or data.get("notes") or ""
    return {
        "type": "entrada",
        "date": _iso_date_only(
            data.get("vencimento") or data.get("dataEmissao") or data.get("date") or data.get("data")
        ),
        "amount": abs(_to_number(data.get("valor") if data.get("valor") is not None else data.get("amount"))),
        "category": (data.get("categoria") or data.get("category") or "conta_receber").strip(),
        "notes": (notes or "").strip(),
        "cliente": (data.get("cliente") or "").strip() or None,
        "notaFiscal": (data.get("notaFiscal") or "").strip() or None,
        "dataEmissao": _iso_date_only(data.get("dataEmissao")) or None,
        "taxasJuros": _to_number(data.get("taxasJuros")),
        "documentoRecebimento": (data.get("documentoRecebimento") or "").strip() or None,
        "dataBaixa": _iso_date_only(data.get("dataBaixa")) or None,
        "valorLiqRecebido": _to_number(data.get("valorLiqRecebido")),
        "action_id": (data.get("action_id") or "").strip() or None,
    }


def _validate(p: Dict[str, Any]) -> Optional[tuple[str, int]]:
    if p.get("type") not in _VALID_TYPES:
        return ("type inválido (entrada, saida, despesa)", 400)
    if not _iso_date_only(p.get("date")):
        return ("date inválido (YYYY-MM-DD)", 400)
    if _to_number(p.get("amount")) <= 0:
        return ("amount deve ser maior que zero", 400)
    return None


# ============================ Rotas base ========================
@finance_bp.route("/transactions", methods=["GET"])
def list_transactions():
    guard = _require_admin()
    if guard:
        return guard
    try:
        ttype = (request.args.get("type") or "").strip().lower()
        action_id = (request.args.get("action_id") or "").strip()
        month = (request.args.get("month") or "").strip()
        return jsonify(_fs_list(ttype, action_id, month)), 200
    except Exception as e:
        log.exception("[finance][GET /transactions] %s", e)
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/transactions", methods=["POST"])
def create_transaction():
    guard = _require_admin()
    if guard:
        return guard
    try:
        data = request.get_json(silent=True) or {}
        payload = {
            "id": str(uuid.uuid4()),
            "type": (data.get("type") or "").strip().lower(),
            "date": _iso_date_only(data.get("date")),
            "amount": abs(_to_number(data.get("amount"))),
            "category": (data.get("category") or "").strip(),
            "notes": (data.get("notes") or "").strip(),
            "action_id": (data.get("action_id") or "").strip() or None,
            "created_by": get_jwt_identity(),
            "created_at": _now_iso(),
        }
        invalid = _validate(payload)
        if invalid:
            msg, code = invalid
            return jsonify({"message": msg}), code

        saved = _fs_create(payload)
        return jsonify(_doc_to_json(saved)), 201
    except Exception as e:
        log.exception("[finance][POST /transactions] %s | body=%s", e, request.get_json(silent=True))
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/transactions/<txid>", methods=["PUT", "PATCH"])
def update_transaction(txid: str):
    guard = _require_admin()
    if guard:
        return guard
    try:
        data = request.get_json(silent=True) or {}
        upd: Dict[str, Any] = {}
        if "type" in data:
            t = str(data["type"]).lower().strip()
            if t in _VALID_TYPES:
                upd["type"] = t
        if any(k in data for k in ("date", "vencimento", "data", "dataEmissao")):
            d = _iso_date_only(
                data.get("date") or data.get("vencimento") or data.get("data") or data.get("dataEmissao")
            )
            if d:
                upd["date"] = d
        if "amount" in data or "valor" in data:
            upd["amount"] = abs(_to_number(data.get("amount") if "amount" in data else data.get("valor")))
        for k_src, k_dst in (("category", "category"), ("categoria", "category")):
            if k_src in data:
                upd[k_dst] = (data.get(k_src) or "").strip()
        if any(k in data for k in ("notes", "descricao", "cliente", "notaFiscal")):
            upd["notes"] = (
                data.get("notes") or data.get("descricao") or data.get("cliente") or data.get("notaFiscal") or ""
            ).strip()

        if not upd:
            return jsonify({"message": "Nada para atualizar"}), 400

        upd["updated_at"] = _now_iso()
        out = _fs_update(txid, upd)
        if not out:
            return jsonify({"message": "Transação não encontrada"}), 404
        return jsonify(_doc_to_json(out)), 200
    except Exception as e:
        log.exception("[finance][PUT /transactions/%s] %s | body=%s", txid, e, request.get_json(silent=True))
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/transactions/<txid>", methods=["DELETE"])
def delete_transaction(txid: str):
    guard = _require_admin()
    if guard:
        return guard
    try:
        _fs_delete(txid)
        return ("", 204)
    except Exception as e:
        log.exception("[finance][DELETE /transactions/%s] %s", txid, e)
        return jsonify({"message": "internal_error"}), 500


# ============================ Aliases pagar/receber ===============
@finance_bp.route("/contas-pagar", methods=["GET"])
def listar_contas_pagar():
    guard = _require_admin()
    if guard:
        return guard
    try:
        ttype = "saida"
        action_id = (request.args.get("action_id") or "").strip()
        month = (request.args.get("month") or "").strip()
        items = _fs_list(ttype, action_id, month)
        return jsonify(items), 200
    except Exception as e:
        log.exception("[finance][GET /contas-pagar] %s", e)
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-receber", methods=["GET"])
def listar_contas_receber():
    guard = _require_admin()
    if guard:
        return guard
    try:
        ttype = "entrada"
        action_id = (request.args.get("action_id") or "").strip()
        month = (request.args.get("month") or "").strip()
        items = _fs_list(ttype, action_id, month)
        return jsonify(items), 200
    except Exception as e:
        log.exception("[finance][GET /contas-receber] %s", e)
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-pagar", methods=["POST"])
def criar_conta_pagar():
    guard = _require_admin()
    if guard:
        return guard
    try:
        data = request.get_json(silent=True) or {}
        mapped = _map_pagar(data)
        if not mapped.get("date"):
            return jsonify({"message": "vencimento/date inválido (YYYY-MM-DD)"}), 400
        if _to_number(mapped.get("amount")) <= 0:
            return jsonify({"message": "valor/amount deve ser maior que zero"}), 400
        doc = {
            "id": str(uuid.uuid4()),
            **mapped,
            "created_by": get_jwt_identity(),
            "created_at": _now_iso(),
        }
        saved = _fs_create(doc)
        return jsonify(_doc_to_json(saved)), 201
    except Exception as e:
        log.exception("[finance][POST /contas-pagar] %s | body=%s", e, request.get_json(silent=True))
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-receber", methods=["POST"])
def criar_conta_receber():
    guard = _require_admin()
    if guard:
        return guard
    try:
        data = request.get_json(silent=True) or {}
        mapped = _map_receber(data)
        if not mapped.get("date"):
            return jsonify({"message": "vencimento/date inválido (YYYY-MM-DD)"}), 400
        if _to_number(mapped.get("amount")) <= 0:
            return jsonify({"message": "valor/amount deve ser maior que zero"}), 400
        doc = {
            "id": str(uuid.uuid4()),
            **mapped,
            "created_by": get_jwt_identity(),
            "created_at": _now_iso(),
        }
        saved = _fs_create(doc)
        return jsonify(_doc_to_json(saved)), 201
    except Exception as e:
        log.exception("[finance][POST /contas-receber] %s | body=%s", e, request.get_json(silent=True))
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-pagar/<txid>", methods=["PUT", "PATCH"])
def atualizar_conta_pagar(txid: str):
    guard = _require_admin()
    if guard:
        return guard
    try:
        data = request.get_json(silent=True) or {}
        upd = _map_pagar(data)
        upd = {k: v for k, v in upd.items() if k in data or k == "type"}
        if "date" in upd and not upd["date"]:
            upd.pop("date")
        upd["updated_at"] = _now_iso()
        out = _fs_update(txid, upd)
        if not out:
            return jsonify({"message": "Transação não encontrada"}), 404
        return jsonify(_doc_to_json(out)), 200
    except Exception as e:
        log.exception("[finance][PUT /contas-pagar/%s] %s | body=%s", txid, e, request.get_json(silent=True))
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-receber/<txid>", methods=["PUT", "PATCH"])
def atualizar_conta_receber(txid: str):
    guard = _require_admin()
    if guard:
        return guard
    try:
        data = request.get_json(silent=True) or {}
        upd = _map_receber(data)
        upd = {k: v for k, v in upd.items() if k in data or k == "type"}
        if "date" in upd and not upd["date"]:
            upd.pop("date")
        upd["updated_at"] = _now_iso()
        out = _fs_update(txid, upd)
        if not out:
            return jsonify({"message": "Transação não encontrada"}), 404
        return jsonify(_doc_to_json(out)), 200
    except Exception as e:
        log.exception("[finance][PUT /contas-receber/%s] %s | body=%s", txid, e, request.get_json(silent=True))
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-pagar/<txid>", methods=["DELETE"])
def deletar_conta_pagar(txid: str):
    guard = _require_admin()
    if guard:
        return guard
    try:
        _fs_delete(txid)
        return ("", 204)
    except Exception as e:
        log.exception("[finance][DELETE /contas-pagar/%s] %s", txid, e)
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-receber/<txid>", methods=["DELETE"])
def deletar_conta_receber(txid: str):
    guard = _require_admin()
    if guard:
        return guard
    try:
        _fs_delete(txid)
        return ("", 204)
    except Exception as e:
        log.exception("[finance][DELETE /contas-receber/%s] %s", txid, e)
        return jsonify({"message": "internal_error"}), 500


# ============================ IMPORT XLSX =======================
def _xlsx_to_iso(v):
    if not v:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return _iso_date_only(v)


@finance_bp.route("/contas-pagar/import", methods=["POST"])
def import_contas_pagar():
    guard = _require_admin()
    if guard:
        return guard
    try:
        if "file" not in request.files:
            return jsonify({"message": "arquivo não enviado (campo 'file')"}), 400
        f = request.files["file"]
        fname = secure_filename(f.filename or "")
        if not fname.lower().endswith((".xlsx", ".xlsm")):
            return jsonify({"message": "formato inválido (use .xlsx)"}), 400

        from openpyxl import load_workbook  # lazy import

        wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
        ws = wb.active

        headers = [str((ws.cell(row=1, column=c).value or "")).strip().lower() for c in range(1, ws.max_column + 1)]

        def col(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        c_venc = col("vencimento")
        c_doc = col("documento")
        c_desc = col("descrição") or col("descricao")
        c_val = col("valor")
        c_dtpg = col("data pagamento") or col("data_pagamento")
        c_valpg = col("valor pago") or col("valorpago")

        if not (c_venc and c_desc and c_val):
            return jsonify({"message": "Cabeçalhos obrigatórios: Vencimento, Descrição, Valor"}), 400

        created = 0
        for r in range(2, ws.max_row + 1):
            venc = ws.cell(row=r, column=c_venc).value
            doc = ws.cell(row=r, column=c_doc).value if c_doc else None
            desc = ws.cell(row=r, column=c_desc).value
            val = ws.cell(row=r, column=c_val).value
            dtpg = ws.cell(row=r, column=c_dtpg).value if c_dtpg else None
            valp = ws.cell(row=r, column=c_valpg).value if c_valpg else None

            payload = {
                "id": str(uuid.uuid4()),
                "type": "saida",
                "date": _xlsx_to_iso(venc),
                "amount": _to_number(val),
                "category": "conta_pagar",
                "notes": (str(desc or "")).strip(),
                "documento": (str(doc or "")).strip() or None,
                "dataPagamento": _xlsx_to_iso(dtpg) or None,
                "valorPago": _to_number(valp),
                "created_by": get_jwt_identity(),
                "created_at": _now_iso(),
            }
            if not payload["date"] or payload["amount"] <= 0:
                continue
            _fs_create(payload)
            created += 1

        return jsonify({"imported": created}), 201
    except Exception as e:
        log.exception("[finance][IMPORT pagar] %s", e)
        return jsonify({"message": "internal_error"}), 500


@finance_bp.route("/contas-receber/import", methods=["POST"])
def import_contas_receber():
    guard = _require_admin()
    if guard:
        return guard
    try:
        if "file" not in request.files:
            return jsonify({"message": "arquivo não enviado (campo 'file')"}), 400
        f = request.files["file"]
        fname = secure_filename(f.filename or "")
        if not fname.lower().endswith((".xlsx", ".xlsm")):
            return jsonify({"message": "formato inválido (use .xlsx)"}), 400

        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
        ws = wb.active

        headers = [str((ws.cell(row=1, column=c).value or "")).strip().lower() for c in range(1, ws.max_column + 1)]

        def col(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        c_venc = col("vencimento")
        c_cli = col("cliente")
        c_nf = col("nota fiscal") or col("notafiscal")
        c_dtem = col("data emissão") or col("data emissao") or col("data_emissao")
        c_val = col("valor")
        c_tx = col("taxas/juros") or col("taxas") or col("juros")
        c_doc = col("doc. recebimento") or col("documento recebimento") or col("documento_recebimento")
        c_baixa = col("data baixa") or col("data_baixa")
        c_liq = col("valor líq. recebido") or col("valor liquido recebido") or col("valor liq recebido")

        if not (c_venc and c_cli and c_val):
            return jsonify({"message": "Cabeçalhos obrigatórios: Vencimento, Cliente, Valor"}), 400

        created = 0
        for r in range(2, ws.max_row + 1):
            venc = ws.cell(row=r, column=c_venc).value
            cli = ws.cell(row=r, column=c_cli).value
            nf = ws.cell(row=r, column=c_nf).value if c_nf else None
            dtem = ws.cell(row=r, column=c_dtem).value if c_dtem else None
            val = ws.cell(row=r, column=c_val).value
            tx = ws.cell(row=r, column=c_tx).value if c_tx else None
            doc = ws.cell(row=r, column=c_doc).value if c_doc else None
            baixa = ws.cell(row=r, column=c_baixa).value if c_baixa else None
            liq = ws.cell(row=r, column=c_liq).value if c_liq else None

            payload = {
                "id": str(uuid.uuid4()),
                "type": "entrada",
                "date": _xlsx_to_iso(venc),
                "amount": _to_number(val),
                "category": "conta_receber",
                "notes": (str(cli or "")).strip(),
                "cliente": (str(cli or "")).strip(),
                "notaFiscal": (str(nf or "")).strip() or None,
                "dataEmissao": _xlsx_to_iso(dtem) or None,
                "taxasJuros": _to_number(tx),
                "documentoRecebimento": (str(doc or "")).strip() or None,
                "dataBaixa": _xlsx_to_iso(baixa) or None,
                "valorLiqRecebido": _to_number(liq),
                "created_by": get_jwt_identity(),
                "created_at": _now_iso(),
            }
            if not payload["date"] or payload["amount"] <= 0:
                continue
            _fs_create(payload)
            created += 1

        return jsonify({"imported": created}), 201
    except Exception as e:
        log.exception("[finance][IMPORT receber] %s", e)
        return jsonify({"message": "internal_error"}), 500
